from UserAPP import Parcelle_BE
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# Call a Workbook() function of openpyxl
# to create a new blank Workbook object


import psycopg2

conn = psycopg2.connect(user="postgres", password="1234", database="xml", host="localhost")
cur = conn.cursor()
cur.execute("SELECT id_parcelle,parcelle_fr,parcelle_ar FROM parcelles order by id_parcelle ASC ;")
reslt = cur.fetchall()

wb = openpyxl.Workbook()
sheet = wb.active  #Get workbook active sheet
sheet.title = "Parcelles"

colColor = openpyxl.styles.colors.Color(rgb='F9F7B8')
colColor2 = openpyxl.styles.colors.Color(rgb='FFFFFF')
columns = ['', 'Numéro', 'Nom_Prop', 'Nom_Prop_ar', 'CIN', 'Douar', 'Douar(ar)', 'Nom Plle(F)', 'Nom Plle(A)', 'Mode', 'Opposition', 'Consistance', 'Surface', 'Mappe']

for col in range(1, len(columns)):
    sheet.cell(1, col).value = columns[col]
    sheet.cell(1, col).font = Font(name='Calibri', size=11, bold=False, italic=False,  vertAlign=None, underline='none', strike=False, color='000000')
    sheet.cell(1, col).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor)
    sheet.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    i = 2
    for n in range(len(reslt)):
        sheet.cell(i, 1).value = reslt[n][0]
        sheet.cell(i, 2).value = reslt[n][1]
        sheet.cell(i, 3).value = reslt[n][2]
        sheet.row_dimensions[i].height = 40
        i += 1

for c in range(1, len(columns)):
    for l in range(2, len(reslt)+2):
        sheet.cell(l, c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor2)
        sheet.cell(l, c).alignment = Alignment(horizontal="center", vertical="center")


sheet.row_dimensions[1].height = 50
sheet.column_dimensions['A'].width = 10
sheet.column_dimensions['B'].width = 32
sheet.column_dimensions['C'].width = 32
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
sheet.column_dimensions['H'].width = 20
sheet.column_dimensions['I'].width = 10
sheet.column_dimensions['J'].width = 12
sheet.column_dimensions['K'].width = 12
sheet.column_dimensions['L'].width = 15
sheet.column_dimensions['M'].width = 12
wb.save('polyg.xlsx')




'''import xlwt
from datetime import datetime

style0 = xlwt.easyxf('font: name Times New Roman , color-index blue, bold on', num_format_str='#,##0.00')
#style1 = xlwt.easyxf(num_format_str='D-MMM-YY')

wb = xlwt.Workbook()
ws = wb.add_sheet('A Test Sheet')

ws.write(0, 0, 1, style0)
ws.write(1, 0, 'laila', style0)

wb.save('test8.xls')'''