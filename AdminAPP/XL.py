import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from UserAPP import Douar_BE
from UserAPP import Parcelle_BE
from UserAPP import Personnes_BE
from UserAPP import Consistance_BE
from UserAPP import Cles
from UserAPP import TypeSol_BE
from UserAPP import TypeSpecl_BE
import os
from tkinter import messagebox

class Exl_Polygs():
    def exl(self):
        wb = openpyxl.Workbook()
        sheet = wb.active  # Get workbook active sheet
        sheet.title = "Parcelles"
        polyg = Parcelle_BE.Parcelle.selectAll(self)

        colColor = openpyxl.styles.colors.Color(rgb='F9F7B8')
        colColor2 = openpyxl.styles.colors.Color(rgb='FFFFFF')
        columns = ['', 'Id_Parcelle', 'Nom Plle(F)', 'Nom Plle(A)', 'Douar', 'Douar(ar)', 'Nom_Prop', 'Nom_Prop_ar', 'Adresse(F)', 'Adresse(A)', 'CIN', 'Tel', 'Opposition', 'Mode', 'Consistance Matérielle', 'Type de Speculation', 'Type de Sol', 'Mappe', 'Surface', 'X', 'Y']


        for col in range(1, len(columns)):
            sheet.cell(1, col).value = columns[col]
            sheet.cell(1, col).font = Font(name='Calibri', size=11, bold=False, italic=False,  vertAlign=None, underline='none', strike=False, color='000000')
            sheet.cell(1, col).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor)
            sheet.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

            lgn = 2
            for v in range(len(polyg)):
                dr = Douar_BE.Douar.xmlDr(self, str(polyg[v][0]))
                person = Personnes_BE.Personnes.prsm_xml(self, str(polyg[v][0]))
                const = Consistance_BE.Consistance.Const_xml(self, str(polyg[v][0]))
                sol = TypeSol_BE.TypeSol.sol_xml(self, str(polyg[v][0]))
                specl = TypeSpecl_BE.TypeSpecul.spec_xml(self, str(polyg[v][0]))

                oppos = Cles.Cles.opposition_pieces(self, str(polyg[v][0]))
                sheet.cell(lgn, 1).value = polyg[v][0] #id_parcel
                sheet.cell(lgn, 2).value = polyg[v][1] #nom_parcel
                sheet.cell(lgn, 3).value = polyg[v][2] #nom_parcel_ar
                sheet.cell(lgn, 4).value = dr[0][1] #adres
                sheet.cell(lgn, 5).value = dr[0][2] #adres_ar
                sheet.cell(lgn, 6).value = person[0][1] +' '+ person[0][3] #prop
                sheet.cell(lgn, 7).value = person[0][2] +' '+ person[0][4] #prop_ar
                sheet.cell(lgn, 8).value = person[0][5] #adres
                sheet.cell(lgn, 9).value = person[0][6] #adres_ar
                sheet.cell(lgn, 10).value = person[0][7] #CIN
                sheet.cell(lgn, 11).value = person[0][8] #Tel
                if (oppos[0]) > 0:
                    sheet.cell(lgn, 12).value = 'O'  #OPPOSITION
                elif (oppos[0]) == 0:
                    sheet.cell(lgn, 12).value = 'N'  #Non OPPOSITION
                sheet.cell(lgn, 13).value = polyg[v][7] #mode
                sheet.cell(lgn, 14).value = const[0][2] #consistance
                sheet.cell(lgn, 15).value = specl[0][2]#specul
                sheet.cell(lgn, 16).value = sol[0][2]#sol
                sheet.cell(lgn, 17).value = polyg[v][16]#mappe
                sheet.cell(lgn, 18).value = ''#surface
                sheet.cell(lgn, 19).value = polyg[v][19]#x
                sheet.cell(lgn, 20).value = polyg[v][20]#y
                lgn += 1

        for c in range(1, len(columns)):
            for l in range(2, len(polyg)+2):
                sheet.cell(l, c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor2)
                sheet.cell(l, c).alignment = Alignment(horizontal="center", vertical="center")


        sheet.row_dimensions[1].height = 50
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 13
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 16
        sheet.column_dimensions['I'].width = 16
        sheet.column_dimensions['J'].width = 8
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 12
        sheet.column_dimensions['M'].width = 12
        sheet.column_dimensions['N'].width = 20
        sheet.column_dimensions['O'].width = 20
        sheet.column_dimensions['P'].width = 12
        sheet.column_dimensions['Q'].width = 10
        sheet.column_dimensions['R'].width = 15
        sheet.column_dimensions['S'].width = 10
        sheet.column_dimensions['T'].width = 10
        try:
            filename = 'Toutes les parcelles.xlsx'
            filepath = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\EXCEL', filename)
            if not os.path.exists(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\EXCEL')):
                os.makedirs(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\EXCEL'))
            wb.save(filepath)
        except:
            msg = messagebox.showerror("Excel", "Veuillez fermer le fichier")

#m = Exl_Polygs()
#m.exl()