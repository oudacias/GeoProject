import tkinter as tk
from tkinter import *
from tkinter import ttk
import os
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from tkcalendar import Calendar, DateEntry
from UserAPP import Parcelle_BE
from UserAPP import SousZone_BE
from UserAPP import Personnes_BE
from UserAPP import Douar_BE
from AdminAPP import User_BE
from AdminAPP import Exportation_BE
from UserAPP import Cles
from UserAPP import TypeSol_BE
from UserAPP import TypeSpecl_BE
from UserAPP import Consistance_BE
from UserAPP import Autocompletecombox
from AdminAPP import ExcelFile

class ExportXls(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.title("Filtrer les parcelles")
        self.geometry("1040x580+160+50")
        self.resizable(0, 0)
        self.config(bg="#F7F9F9")
        #self.iconbitmap(r'C:\Users\LAILA\PycharmProjects\IFE_GIS\icons\iconlogo.ico')


#FILTRER LES PARCELLES
        lf = LabelFrame(self, text="Filtrer les parcelles", width=348, height=502, font=("Helvetica", 12), bg="#F7F9F9")
        lf.place(x=4, y=4)

        super_id = Label(lf, text="Id supérieur ou égal à ", bg="#F7F9F9", font=("Helvetica", 11))
        super_id.place(x=10, y=15)
        polyg = Parcelle_BE.Parcelle()
        idPolyg = polyg.combboxID()
        self.super_id_box = ttk.Combobox(lf, values=idPolyg, width=25)
        self.super_id_box.place(x=160, y=15)

        infr_id = Label(lf, text="Id inférieur ou égal à ", bg="#F7F9F9", font=("Helvetica", 11))
        infr_id.place(x=10, y=55)
        self.infrId_box = ttk.Combobox(lf, values=idPolyg, width=25)
        self.infrId_box.place(x=160, y=55)

        date_inferieur = Label(lf, text="Date _De", bg="#F7F9F9", font=("Helvetica", 11))
        date_inferieur.place(x=47, y=95)
        self.infrDt = DateEntry(lf, date_pattern='dd-mm-yyyy', width=25)
        self.infrDt.place(x=160, y=95)
        self.infrDt.config(selectbackground='gray80',
                           selectforeground='black',
                           normalbackground='white',
                           normalforeground='black',
                           background='gray90',
                           foreground='black',
                           bordercolor='gray90',
                           othermonthforeground='gray50',
                           othermonthbackground='white',
                           othermonthweforeground='gray50',
                           othermonthwebackground='white',
                           weekendbackground='white',
                           weekendforeground='black',
                           headersbackground='white',
                           headersforeground='gray70')
        self.infrDt.delete(0, 'end')

        date_superieur = Label(lf, text="Date _A:", bg="#F7F9F9", font=("Helvetica", 11))
        date_superieur.place(x=47, y=135)
        self.superDt = DateEntry(lf, date_pattern='dd-mm-yyyy', width=25)
        self.superDt.place(x=160, y=135)
        self.superDt.config(selectbackground='gray80',
                            selectforeground='black',
                            normalbackground='white',
                            normalforeground='black',
                            background='gray90',
                            foreground='black',
                            bordercolor='gray90',
                            othermonthforeground='gray50',
                            othermonthbackground='white',
                            othermonthweforeground='gray50',
                            othermonthwebackground='white',
                            weekendbackground='white',
                            weekendforeground='black',
                            headersbackground='white',
                            headersforeground='gray70')
        self.superDt.delete(0, 'end')

        ss_zone = Label(lf, text="Sous-zone ", bg="#F7F9F9", font=("Helvetica", 11))
        ss_zone.place(x=46, y=175)
        sz = SousZone_BE.SousZ()
        szval = sz.combboxSZ()
        self.sZn_box = Autocompletecombox.Autocompletecombox(lf, values=szval, width=25)
        self.sZn_box.set_completion_list(szval)
        self.sZn_box.place(x=160, y=175)

        idDouar = Label(lf, text="Douar ", bg="#F7F9F9", font=("Helvetica", 11))
        idDouar.place(x=49, y=215)
        dr = Douar_BE.Douar()
        drVal = dr.combboxcDr()
        self.dr_box = Autocompletecombox.Autocompletecombox(lf, values=drVal, width=25)
        self.dr_box.set_completion_list(drVal)
        self.dr_box.place(x=160, y=215)

        cin = Label(lf, text="CIN ", bg="#F7F9F9", font=("Helvetica", 11))
        cin.place(x=58, y=255)
        prs = Personnes_BE.Personnes()
        cinVal = prs.combboxcCIN()
        self.cin_box = Autocompletecombox.Autocompletecombox(lf, values=cinVal, width=25)
        self.cin_box.set_completion_list(cinVal)
        self.cin_box.place(x=160, y=255)

        nmUsr = Label(lf, text="Utilisateur ", bg="#F7F9F9", font=("Helvetica", 11))
        nmUsr.place(x=45, y=295)
        usr = User_BE.User()
        usrVal = usr.usersList()
        self.nmUsr_box = Autocompletecombox.Autocompletecombox(lf, values=usrVal, width=25)
        self.nmUsr_box.set_completion_list(usrVal)
        self.nmUsr_box.place(x=160, y=295)

        listP = Label(lf, text="Parcelles Nº ", bg="#F7F9F9", font=("Helvetica", 11))
        listP.place(x=32, y=335)
        self.lst = Text(lf, relief=FLAT)
        self.lst.place(x=160, y=335, width=170, height=80)

        filterButton = Button(lf, text="Filtrer", font=("Times New Roman", 11), bg="#4EB1FA", fg="#FFFFFF", relief=FLAT, activebackground="#9FE5D7", activeforeground="#E8B24D", command=self.fil)
        filterButton.place(x=150, y=440)


#LA LISTE A EXPORTER
        style = ttk.Style(self)
        style.theme_use('clam')
        style.configure("Treeview", background="#FFFFFF", foreground="#14566D")
        style.configure("Treeview.Heading", background="#B0D9E8", foreground="#14566D")

        lf = LabelFrame(self, text="La liste a exporter", font=("Helvetica", 12), bg="#F7F9F9", width=682, height=502)
        lf.place(x=355, y=4)

        self.tree = ttk.Treeview(lf)
        self.tree.place(x=3, y=6, width=657, height=468)

        ysb = ttk.Scrollbar(self, orient='vertical', command=self.tree.yview)
        ysb.place(x=1017, y=32, height=468)

        xsb = ttk.Scrollbar(self, orient='horizontal', command=self.tree.xview)
        xsb.place(x=360, y=486, width=657)
        self.tree['yscroll'] = ysb.set
        self.tree['xscroll'] = xsb.set

        self.parcelle = Parcelle_BE.Parcelle()
        self.col_name = self.parcelle.columnName()
        self.tree["columns"] = ("0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15")
        self.tree['show'] = 'headings'
        for c in range(len(self.tree["columns"])):
            self.tree.column(c, width=100)
            self.tree.heading(c, text=self.col_name[c])

# EXPORTER
        lfBtn = LabelFrame(self, text="Exporter", font=("Helvetica", 12), bg="#F7F9F9", width=1031, height=70)
        lfBtn.place(x=5, y=505)
        frm2 = Frame(lfBtn, bg="#379BF3", width=169, height=26)
        frm2.place(x=420, y=8)
        self.btn1 = tk.Button(frm2, text="Exporter Vers EXCEL", bg="#CEE6F3", font=("Times New Roman", 13), fg="#379BF3", relief='flat', activebackground="#CEE6F3", activeforeground="#379BF3", command=self.export)
        self.btn1.place(x=1, y=1, height=24)

    def fil(self):
        filtr = Exportation_BE.Filtrage()
        self.list = filtr.filtrage(self.super_id_box.get(), self.infrId_box.get(), self.infrDt.get(), self.superDt.get(), self.sZn_box.get(), self.dr_box.get(), self.cin_box.get(), self.nmUsr_box.get(), self.lst.get('1.0', 'end-1c'))

# vider les champs
        self.super_id_box.delete(0, 'end')
        self.infrId_box.delete(0, 'end')
        self.infrDt.delete(0, 'end')
        self.superDt.delete(0, 'end')
        self.sZn_box.delete(0, 'end')
        self.dr_box.delete(0, 'end')
        self.cin_box.delete(0, 'end')
        self.nmUsr_box.delete(0, 'end')
        self.lst.delete(1.0, 'end')

# insert values
        self.tree.delete(*self.tree.get_children())
        index = iid = 0
        for self.item in self.list:
            self.tree.insert("", index, iid, values=self.item)
            index = iid = index + 1
        self.tree.bind("<Double-1>", self.OnDoubleClick)

    def OnDoubleClick(self, event):
        self.root = Toplevel()
        self.root.geometry("240x140+600+260")
        self.root.resizable(0, 0)
        self.root.config(bg="#F7F9F9")
        self.root.title("Excel")
        lbl = Label(self.root, text="Voulez-Vous retirer cette ligne \n de la liste?", font="Helvetica", bg="#F7F9F9")
        lbl.place(x=11, y=18)
        btn = Button(self.root, text="Retirer", command=self.retirer, bg="#4EB1FA", fg="#FFFFFF", relief=FLAT, activebackground="#4EB1FA", activeforeground="#000000")
        btn.place(x=70, y=80)
        btn2 = Button(self.root, text="Annuler", command=self.annuler, fg="#FFFFFF", bg="#FF862E", relief=FLAT, activebackground="#FF8A02", activeforeground="#000000")
        btn2.place(x=135, y=80)

    def retirer(self):
        x = self.tree.selection()[0]
        self.tree.delete(x)
        self.root.destroy()

    def annuler(self):
        self.root.destroy()

    def export(self):

        wb = openpyxl.Workbook()
        sheet = wb.active  # Get workbook active sheet
        sheet.title = "Parcelles"
        columns = ['', 'Id_Parcelle', 'Nom Plle(F)', 'Nom Plle(A)', 'Douar', 'Douar(ar)', 'Nom_Prop', 'Nom_Prop_ar',
                   'Adresse(F)', 'Adresse(A)', 'CIN', 'Tel', 'Opposition', 'Mode', 'Consistance Matérielle',
                   'Type de Speculation', 'Type de Sol', 'Mappe', 'Surface', 'X', 'Y']
        colColor = openpyxl.styles.colors.Color(rgb='F9F7B8')
        colColor2 = openpyxl.styles.colors.Color(rgb='FFFFFF')

        polyg = Parcelle_BE.Parcelle.selectAll(self)
        for col in range(1, len(columns)):
            sheet.cell(1, col).value = columns[col]
            sheet.cell(1, col).font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None,
                                           underline='none', strike=False, color='000000')
            sheet.cell(1, col).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor)
            sheet.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

            lgn = 2
            for self.item in range(len(self.list)):
                dr = Douar_BE.Douar.xmlDr(self, str(polyg[self.item][0]))
                person = Personnes_BE.Personnes.prsm_xml(self, str(polyg[self.item][0]))
                const = Consistance_BE.Consistance.Const_xml(self, str(polyg[self.item][0]))
                sol = TypeSol_BE.TypeSol.sol_xml(self, str(polyg[self.item][0]))
                specl = TypeSpecl_BE.TypeSpecul.spec_xml(self, str(polyg[self.item][0]))

                oppos = Cles.Cles.opposition_pieces(self, str(polyg[self.item][0]))
                sheet.cell(lgn, 1).value = polyg[self.item][0]  # id_parcel
                sheet.cell(lgn, 2).value = polyg[self.item][1]  # nom_parcel
                sheet.cell(lgn, 3).value = polyg[self.item][2]  # nom_parcel_ar
                sheet.cell(lgn, 4).value = dr[0][1]  # adres
                sheet.cell(lgn, 5).value = dr[0][2]  # adres_ar
                sheet.cell(lgn, 6).value = person[0][1] + ' ' + person[0][3]  # prop
                sheet.cell(lgn, 7).value = person[0][2] + ' ' + person[0][4]  # prop_ar
                sheet.cell(lgn, 8).value = person[0][5]  # adres
                sheet.cell(lgn, 9).value = person[0][6]  # adres_ar
                sheet.cell(lgn, 10).value = person[0][7]  # CIN
                sheet.cell(lgn, 11).value = person[0][8]  # Tel
                if (oppos[0]) > 0:
                    sheet.cell(lgn, 12).value = 'O'  # OPPOSITION
                elif (oppos[0]) == 0:
                    sheet.cell(lgn, 12).value = 'N'  # Non OPPOSITION
                sheet.cell(lgn, 13).value = polyg[self.item][7]  # mode
                sheet.cell(lgn, 14).value = const[0][2]  # consistance
                sheet.cell(lgn, 15).value = specl[0][2]  # specul
                sheet.cell(lgn, 16).value = sol[0][2]  # sol
                sheet.cell(lgn, 17).value = ''  # mappe
                sheet.cell(lgn, 18).value = ''  # surface
                sheet.cell(lgn, 19).value = polyg[self.item][19]  # x
                sheet.cell(lgn, 20).value = polyg[self.item][20]  # y
                lgn += 1

        for c in range(1, len(columns)):
            for l in range(2, len(polyg)+2):
                sheet.cell(l, c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor2)
                sheet.cell(l, c).alignment = Alignment(horizontal="center", vertical="center")

        #for self.item in enumerate(self.list):
        #print(max(str(self.item[0])),  min(str(self.item[0])))
        #for self.item in self.list:
            #print(self.item[0])

        sheet.row_dimensions[1].height = 50
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 13
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 16
        sheet.column_dimensions['I'].width = 16
        sheet.column_dimensions['J'].width = 8
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 8
        sheet.column_dimensions['N'].width = 25
        sheet.column_dimensions['O'].width = 25
        sheet.column_dimensions['P'].width = 12
        sheet.column_dimensions['Q'].width = 10
        sheet.column_dimensions['R'].width = 15
        sheet.column_dimensions['S'].width = 10
        sheet.column_dimensions['T'].width = 10
        try:
            x = len(self.list)
            id1 = str(self.list[0][0])
            id2 = str(self.list[x - 1][0])
            if x == 1:
                filename = 'P' + id1 +'.xlsx'
                filepath = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\EXCEL', filename)

                if not os.path.exists(os.path.join(os.path.join(os.path.expanduser('~')),  'Desktop\IFE_PIECES\EXCEL')):
                    os.makedirs(os.path.join(os.path.join(os.path.expanduser('~')),  'Desktop\IFE_PIECES\EXCEL'))
                    
                wb.save(filepath)
                self.tree.delete(*self.tree.get_children())
                ExcelFile.XL_File()
            elif x > 1:
                filename = 'P'+id1+'...P'+id2+'.xlsx'
                filepath = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\EXCEL', filename)
                if not os.path.exists(os.path.join(os.path.join(os.path.expanduser('~')),  'Desktop\IFE_PIECES\EXCEL')):
                    os.makedirs(os.path.join(os.path.join(os.path.expanduser('~')),  'Desktop\IFE_PIECES\EXCEL'))
                wb.save(filepath)
                self.tree.delete(*self.tree.get_children())
                ExcelFile.XL_File()
        except:
            msg = messagebox.showerror("Excel", "Veuillez fermer le fichier")

'''if __name__ == "__main__":
    app = ExportXls()
    app.mainloop()'''
