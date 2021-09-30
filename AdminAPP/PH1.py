import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from UserAPP import Douar_BE
from UserAPP import Parcelle_BE
from UserAPP import Personnes_BE
from UserAPP import Consistance_BE
from UserAPP import TypeSol_BE
from UserAPP import TypeSpecl_BE
from tkinter import messagebox
import os

class Reper_Ph1():

    def ph1(self):
        wb = openpyxl.Workbook()
        sheet = wb.active  # Get workbook active sheet
        sheet.title = "PH1"
        polyg = Parcelle_BE.Parcelle.selectAll(self)

        colColor = openpyxl.styles.colors.Color(rgb='F9F7B8')
        colColor2 = openpyxl.styles.colors.Color(rgb='FFFFFF')
        columns = ['', 'Carnet', 'Page', 'Ordre', 'N° plle', 'BIS', 'Numéro Rtion', 'Indice', 'Date de Dépôt', 'Nom Arabe', 'Prénom Arabe', 'Autre Nom Arabe',
                   'Nom Français', 'Prénom Français', 'Autre Nom (F)', 'Date\n Naissance', 'CINE', 'Quôte Numérateur', 'Quôte Dénominateur', 'Adresse(F)', 'Adresse(A)',
                   'He', 'A', 'Ca', 'Nom Plle(F)', 'Nom Plle(A)', 'Nature(F)', 'Nature Principale(A)', 'Autres natures (A)', 'Mappe', 'Consistance Matérielle', 'Type de Speculation', 'Type de Sol', 'X', 'Y', 'Observations']


        for col in range(1, len(columns)):
            sheet.cell(1, col).value = columns[col]
            sheet.cell(1, col).font = Font(name='Calibri', size=11, bold=False, italic=False,  vertAlign=None, underline='none', strike=False, color='000000')
            sheet.cell(1, col).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor)
            sheet.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

            lgn = 2
            for v in range(len(polyg)):
                dr = Douar_BE.Douar.xmlDr(self, str(polyg[v][0]))
                person = Personnes_BE.Personnes.prsm_xml(self, str(polyg[v][0]))
                const = Consistance_BE.Consistance.Const_xml(self, str(polyg[v][0]))
                sol = TypeSol_BE.TypeSol.sol_xml(self, str(polyg[v][0]))
                specl = TypeSpecl_BE.TypeSpecul.spec_xml(self, str(polyg[v][0]))

                sheet.cell(lgn, 1).value = ''
                sheet.cell(lgn, 2).value = ''
                sheet.cell(lgn, 3).value = ''
                sheet.cell(lgn, 4).value = polyg[v][0]
                sheet.cell(lgn, 5).value = ''
                sheet.cell(lgn, 6).value = ''
                sheet.cell(lgn, 7).value = ''
                sheet.cell(lgn, 8).value = ''
                sheet.cell(lgn, 9).value = person[0][2]
                sheet.cell(lgn, 10).value = person[0][4]
                sheet.cell(lgn, 11).value = ''
                sheet.cell(lgn, 12).value = person[0][1]
                sheet.cell(lgn, 13).value = person[0][3]
                sheet.cell(lgn, 14).value = ''
                if person[0][10] != '':
                    sheet.cell(lgn, 15).value = person[0][10]
                else:
                    sheet.cell(lgn, 15).value = ''
                sheet.cell(lgn, 16).value = person[0][7]
                sheet.cell(lgn, 17).value = ''
                sheet.cell(lgn, 18).value = ''
                sheet.cell(lgn, 19).value = dr[0][1]
                sheet.cell(lgn, 20).value = dr[0][2]
                sheet.cell(lgn, 21).value = polyg[v][20]
                sheet.cell(lgn, 22).value = polyg[v][21]
                sheet.cell(lgn, 23).value = polyg[v][22]
                sheet.cell(lgn, 24).value = polyg[v][1]
                sheet.cell(lgn, 25).value = polyg[v][2]
                sheet.cell(lgn, 26).value = const[0][2]
                sheet.cell(lgn, 27).value = const[0][3]
                sheet.cell(lgn, 28).value = ''
                sheet.cell(lgn, 29).value = ''
                sheet.cell(lgn, 30).value = const[0][2]
                sheet.cell(lgn, 31).value = specl[0][2]
                sheet.cell(lgn, 32).value = sol[0][2]
                sheet.cell(lgn, 33).value = polyg[v][19]
                sheet.cell(lgn, 34).value = polyg[v][20]
                sheet.cell(lgn, 35).value = '' #OBSERVATION
                sheet.cell(lgn, 36).value = polyg[v][16]
                sheet.row_dimensions[lgn].height = 40
                lgn += 1

        for c in range(1, len(columns)):
            for l in range(2, len(polyg)+2):
                sheet.cell(l, c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor2)
                sheet.cell(l, c).alignment = Alignment(horizontal="center", vertical="center")


        sheet.row_dimensions[1].height = 50
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 20
        sheet.column_dimensions['L'].width = 20
        sheet.column_dimensions['M'].width = 20
        sheet.column_dimensions['N'].width = 20
        sheet.column_dimensions['O'].width = 15
        sheet.column_dimensions['P'].width = 15
        sheet.column_dimensions['Q'].width = 12
        sheet.column_dimensions['R'].width = 12
        sheet.column_dimensions['S'].width = 15
        sheet.column_dimensions['T'].width = 15
        sheet.column_dimensions['U'].width = 8
        sheet.column_dimensions['V'].width = 8
        sheet.column_dimensions['W'].width = 8
        sheet.column_dimensions['X'].width = 20
        sheet.column_dimensions['Y'].width = 20
        sheet.column_dimensions['Z'].width = 10
        sheet.column_dimensions['AA'].width = 20
        sheet.column_dimensions['AB'].width = 20
        sheet.column_dimensions['AC'].width = 12
        sheet.column_dimensions['AD'].width = 20
        sheet.column_dimensions['AE'].width = 18
        sheet.column_dimensions['AF'].width = 15
        sheet.column_dimensions['AG'].width = 18
        sheet.column_dimensions['AH'].width = 18
        sheet.column_dimensions['AI'].width = 20
        try:
            filename = 'Repertoire-PH1.xlsx'
            filepath = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\Repertoire_PH1',filename)
            if not os.path.exists(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\Repertoire_PH1')):
                os.makedirs(os.path.join(os.path.join(os.path.expanduser('~')),'Desktop\IFE_PIECES\Repertoire_PH1'))
            wb.save(filepath)
        except:
            msg = messagebox.showerror("Repertoire-PH1", "Veuillez fermer le fichier")

#m = Reper_Ph1()
#m.ph1()