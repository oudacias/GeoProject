import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from UserAPP import Parcelle_BE
from tkinter import messagebox
from UserAPP import Cles

class Reper_Ph8():
    def ph8(self):
        wb = openpyxl.Workbook()
        sheet = wb.active  # Get workbook active sheet
        sheet.title = "PH8"
        polyg = Parcelle_BE.Parcelle.selectAll(self)

        colColor = openpyxl.styles.colors.Color(rgb='F9F7B8')
        colColor2 = openpyxl.styles.colors.Color(rgb='FFFFFF')
        columns = ['', 'N° de La Parcelle', "N° d'Ordre", 'N° du Carnet', 'N° de la page dans la Carnet', "N° de la page dans l'Etat Parcellaire", 'SURFACE', 'MAPPE', 'X(m)', 'Y(m)', 'Opposition O:N', 'N° de Réquisition', 'N° de Titre']


        for col in range(1, len(columns)):
            sheet.cell(1, col).value = columns[col]
            sheet.cell(1, col).font = Font(name='Calibri', size=11, bold=False, italic=False,  vertAlign=None, underline='none', strike=False, color='000000')
            sheet.cell(1, col).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor)
            sheet.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

            lgn = 2
            for v in range(len(polyg)):
                oppos = Cles.Cles.opposition_pieces(self, str(polyg[v][0]))

                sheet.cell(lgn, 1).value = polyg[v][0]
                sheet.cell(lgn, 2).value = ''
                sheet.cell(lgn, 3).value = ''
                sheet.cell(lgn, 4).value = ''
                sheet.cell(lgn, 5).value = ''
                sheet.cell(lgn, 6).value = str(polyg[v][20])+ 'Ha  ' + str(polyg[v][21])+ 'A ' + str(polyg[v][22]) + 'CA' #surface
                sheet.cell(lgn, 7).value = polyg[v][16] #mappe
                sheet.cell(lgn, 8).value = polyg[v][19] #x
                sheet.cell(lgn, 9).value = polyg[v][20] #y
                if (oppos[0]) > 0:
                    sheet.cell(lgn, 10).value = 'O'  #OPPOSITION
                elif (oppos[0]) == 0:
                    sheet.cell(lgn, 10).value = 'N'  #Non OPPOSITION
                sheet.cell(lgn, 11).value = polyg[v][12] #requisition
                sheet.cell(lgn, 12).value = polyg[v][13] #titre
                sheet.row_dimensions[lgn].height = 40
                lgn += 1

        for c in range(1, len(columns)):
            for l in range(2, len(polyg)+2):
                sheet.cell(l, c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor2)
                sheet.cell(l, c).alignment = Alignment(horizontal="center", vertical="center")


        sheet.row_dimensions[1].height = 50
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 20
        sheet.column_dimensions['L'].width = 20
        try:
            filename = 'Repertoire-PH8.xlsx'
            filepath = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\Repertoire_PH8',filename)
            if not os.path.exists(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\Repertoire_PH8')):
                os.makedirs(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\IFE_PIECES\Repertoire_PH8'))
            wb.save(filepath)
        except:
            msg = messagebox.showerror("Repertoire-PH8", "Veuillez fermer le fichier")
#m = Reper_Ph8()
#m.ph8()