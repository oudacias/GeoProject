import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from UserAPP import Douar_BE
from UserAPP import Parcelle_BE
from UserAPP import Personnes_BE
from UserAPP import Consistance_BE
from UserAPP import Cles
from tkinter import messagebox
import os

class EtatParcellaire():
    def etat_parce(self):
        wb = openpyxl.Workbook()
        sheet = wb.active  # Get workbook active sheet
        sheet.title = "Parcelles"
        polyg = Parcelle_BE.Parcelle.selectAll(self)

        colColor = openpyxl.styles.colors.Color(rgb='F9F7B8')
        colColor2 = openpyxl.styles.colors.Color(rgb='FFFFFF')
        columns = ['', 'Numéro', 'Nom_Prop', 'Nom_Prop_ar', 'CIN', 'Douar', 'Douar(ar)', 'Nom Plle(F)', 'Nom Plle(A)', 'Mode', 'Opposition', 'Consistance', 'Surface', 'Mappe']


        for col in range(1, len(columns)):
            sheet.cell(1, col).value = columns[col]
            sheet.cell(1, col).font = Font(name='Calibri', size=11, bold=False, italic=False,  vertAlign=None, underline='none', strike=False, color='000000')
            sheet.cell(1, col).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor)
            sheet.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

            lgn = 2
            for v in range(len(polyg)):
                dr = Douar_BE.Douar.xmlDr(self, str(polyg[v][0]))
                person = Personnes_BE.Personnes.prsm_xml(self, str(polyg[v][0]))
                const = Consistance_BE.Consistance.Const_xml(self, str(polyg[v][0]))
                oppos = Cles.Cles.opposition_pieces(self, str(polyg[v][0]))
                sheet.cell(lgn, 1).value = polyg[v][0]
                sheet.cell(lgn, 2).value = person[0][1] + ' ' + person[0][3]
                sheet.cell(lgn, 3).value = person[0][2] + ' ' + person[0][4]
                sheet.cell(lgn, 4).value = person[0][7]
                sheet.cell(lgn, 5).value = dr[0][1]
                sheet.cell(lgn, 6).value = dr[0][2]
                sheet.cell(lgn, 7).value = polyg[v][1]
                sheet.cell(lgn, 8).value = polyg[v][2]
                sheet.cell(lgn, 9).value = polyg[v][7]
                if (oppos[0]) > 0:
                    sheet.cell(lgn, 10).value = 'O'  #OPPOSITION
                elif (oppos[0]) == 0:
                    sheet.cell(lgn, 10).value = 'N'  #Non OPPOSITION
                sheet.cell(lgn, 11).value = const[0][2]
                sheet.cell(lgn, 12).value = str(polyg[v][20])+ 'Ha  ' + str(polyg[v][21])+ 'A ' + str(polyg[v][22]) + 'CA' # surface
                sheet.cell(lgn, 13).value = polyg[v][16]
                sheet.row_dimensions[lgn].height = 40
                lgn += 1

        for c in range(1, len(columns)):
            for l in range(2, len(polyg)+2):
                sheet.cell(l, c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=colColor2)
                sheet.cell(l, c).alignment = Alignment(horizontal="center", vertical="center")


        sheet.row_dimensions[1].height = 50
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 32
        sheet.column_dimensions['C'].width = 32
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 12
        sheet.column_dimensions['K'].width = 12
        sheet.column_dimensions['L'].width = 15
        sheet.column_dimensions['M'].width = 12
        try:
            filename = 'Repertoire-EtatParcellaire.xlsx'
            filepath = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\\IFE_PIECES\\Repertoire_EtatParcellaire', filename)
            if not os.path.exists(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\\FE_PIECES\\Repertoire_EtatParcellaire')):
                os.makedirs(os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop\\IFE_PIECES\\Repertoire_EtatParcellaire'))
            wb.save(filepath)
            print(111)
        except:
            msg = messagebox.showerror("EtatParcellaire", "Veuillez fermer le fichier")

#m = EtatParcellaire()
#m.etat_parce()